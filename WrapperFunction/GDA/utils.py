def num_tokens_from_string(string: str, encoding_name: str) -> int:
    import tiktoken

    """Returns the number of tokens in a text string."""
    encoding = tiktoken.get_encoding(encoding_name)
    num_tokens = len(encoding.encode(string))
    return num_tokens


def print_df_info(df):
    print(
        "\n================================================================================\n"
    )
    print("               INFO on Dataframe object: ", "df")
    print(df.describe())
    print(
        "\n__________________________________________\n__________________________________________\n"
    )
    print(df.info())
    print(
        "\n__________________________________________\n__________________________________________\n"
    )


def get_tmp_filename(extension):
    from datetime import datetime

    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d_%H.%M.%S")

    outfilepath = None

    outfilepath = "Received_File_" + current_time + "." + extension
    return outfilepath


# Used to convert all SAS .xlsx files into .csv files, so they can be merged
def excel2csv(in_folder_path):
    import csv
    import os

    from openpyxl import load_workbook, utils

    delimiter = "|"

    # Iterate through all files in the folder
    for file_name in os.listdir(in_folder_path):
        # Check if the file is an Excel file
        if file_name.endswith(".xlsx"):
            print("Loading workbook: ", file_name)
            # Load the workbook
            wb = load_workbook(os.path.join(in_folder_path, file_name))
            # Iterate through all sheets in the workbook
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Create a new CSV file with the same name as the Excel file
                new_csv_filepath = os.path.join(
                    in_folder_path, file_name.replace(".xlsx", ".csv")
                )
                with open(
                    new_csv_filepath, "w", newline="", encoding="utf-8"
                ) as csv_file:
                    csv_writer = csv.writer(
                        csv_file, delimiter=delimiter, quoting=csv.QUOTE_ALL
                    )
                    # Iterate through all rows in the sheet
                    for row in sheet.iter_rows():
                        csv_writer.writerow(
                            [utils.escape.unescape(str(cell.value)) for cell in row]
                        )
                print("CSV file written: ", new_csv_filepath)
                csv_file.close()
                del csv_writer
            wb.close()
            del wb
